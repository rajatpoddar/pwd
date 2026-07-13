from flask import (
    Flask, render_template, request, jsonify,
    send_file, session
)
import json, os, hashlib, io, uuid
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'NregaBot_SuperSecret_Key_2024_RAJAT'

DATA_DIR   = os.path.join(os.path.dirname(__file__), 'data')
USERS_FILE = os.path.join(DATA_DIR, 'users.json')

# ── Admin user ID (aapka fixed admin ID) ──────────────────────────────────────
ADMIN_USER_ID = 'rajat_admin'

SALT = 'NregaBot_salt_RAJAT_2024'

# ── Helpers ───────────────────────────────────────────────────────────────────
def hash_pin(pin: str) -> str:
    return hashlib.sha256((pin + SALT).encode()).hexdigest()

def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)

def load_users() -> dict:
    ensure_data_dir()
    if not os.path.exists(USERS_FILE):
        # Default: admin user seed
        default = {
            ADMIN_USER_ID: {
                'id': ADMIN_USER_ID,
                'name': 'Rajat',
                'block': 'Admin Block',
                'email': 'rajat@admin.com',
                'mobile': '9999999999',
                'pin_hash': hash_pin('1234'),
                'role': 'admin',
                'lock_timeout': 15,
                'data': get_default_data()
            }
        }
        save_users(default)
        return default
    with open(USERS_FILE, encoding='utf-8') as f:
        return json.load(f)

def save_users(users: dict):
    ensure_data_dir()
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

def get_default_data() -> dict:
    return {
        "categories": [
            {
                "id": "nrega",
                "name": "NREGA",
                "columns": ["P.Code","Panchayat","PS User","PS Pass","Mukhiya User","Mukhiya Pass"],
                "entries": []
            }
        ]
    }

def current_user_id():
    return session.get('user_id')

def get_current_user():
    uid = current_user_id()
    if not uid:
        return None
    users = load_users()
    return users.get(uid)

def require_login(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not current_user_id():
            return jsonify({'ok': False, 'msg': 'Login required'}), 401
        return f(*args, **kwargs)
    return wrapper

def require_pin_verified(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('pin_verified'):
            return jsonify({'ok': False, 'msg': 'PIN verification required'}), 403
        return f(*args, **kwargs)
    return wrapper

def require_admin(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user or user.get('role') != 'admin':
            return jsonify({'ok': False, 'msg': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return wrapper

# ── Main route ─────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

# ══════════════════════════════════════════════════════════════════════════════
# AUTH APIs
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/auth/register', methods=['POST'])
def api_register():
    body   = request.get_json(force=True)
    name   = str(body.get('name', '')).strip()
    block  = str(body.get('block', '')).strip()
    email  = str(body.get('email', '')).strip().lower()
    mobile = str(body.get('mobile', '')).strip()
    pin    = str(body.get('pin', '')).strip()

    # Validation
    if not name:
        return jsonify({'ok': False, 'msg': 'Naam zaruri hai!'}), 400
    if not block:
        return jsonify({'ok': False, 'msg': 'Block naam zaruri hai!'}), 400
    if not email or '@' not in email:
        return jsonify({'ok': False, 'msg': 'Valid email dalein!'}), 400
    if not mobile.isdigit() or len(mobile) != 10:
        return jsonify({'ok': False, 'msg': 'Mobile number 10 digit ka hona chahiye!'}), 400
    if not pin.isdigit() or len(pin) != 4:
        return jsonify({'ok': False, 'msg': 'PIN sirf 4 digit ka hona chahiye!'}), 400

    users = load_users()

    # Check duplicate email or mobile
    for u in users.values():
        if u.get('email') == email:
            return jsonify({'ok': False, 'msg': 'Ye email pehle se registered hai!'}), 400
        if u.get('mobile') == mobile:
            return jsonify({'ok': False, 'msg': 'Ye mobile number pehle se registered hai!'}), 400

    # Create user
    uid = 'user_' + uuid.uuid4().hex[:12]
    users[uid] = {
        'id': uid,
        'name': name,
        'block': block,
        'email': email,
        'mobile': mobile,
        'pin_hash': hash_pin(pin),
        'role': 'user',
        'lock_timeout': 15,
        'data': get_default_data()
    }
    save_users(users)
    return jsonify({'ok': True, 'msg': f'Registration successful! Welcome, {name}!'})


@app.route('/api/auth/login', methods=['POST'])
def api_login():
    body       = request.get_json(force=True)
    identifier = str(body.get('identifier', '')).strip().lower()

    if not identifier:
        return jsonify({'ok': False, 'msg': 'Email ya mobile number dalein!'}), 400

    users = load_users()
    matched_user = None
    for u in users.values():
        if u.get('email') == identifier or u.get('mobile') == identifier:
            matched_user = u
            break

    if not matched_user:
        return jsonify({'ok': False, 'msg': 'Koi account nahi mila is email/mobile se!'}), 404

    # Set session — PIN verification still pending
    session['user_id']      = matched_user['id']
    session['pin_verified'] = False
    return jsonify({
        'ok': True,
        'name': matched_user['name'],
        'msg': f'Welcome {matched_user["name"]}! Ab PIN dalein.'
    })


@app.route('/api/auth/pin/verify', methods=['POST'])
@require_login
def api_pin_verify():
    body = request.get_json(force=True)
    pin  = str(body.get('pin', ''))

    users = load_users()
    user  = users.get(current_user_id())
    if not user:
        return jsonify({'ok': False, 'msg': 'User nahi mila!'}), 404

    if hash_pin(pin) == user.get('pin_hash', ''):
        session['pin_verified'] = True
        return jsonify({
            'ok': True,
            'user_id': current_user_id(),
            'name': user['name'],
            'role': user.get('role', 'user'),
            'lock_timeout': user.get('lock_timeout', 15)
        })
    return jsonify({'ok': False, 'msg': 'Galat PIN! Dobara try karein.'}), 401


@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'ok': True})


@app.route('/api/auth/status', methods=['GET'])
def api_auth_status():
    uid = current_user_id()
    if not uid:
        return jsonify({'logged_in': False, 'pin_verified': False})
    users = load_users()
    user  = users.get(uid)
    if not user:
        session.clear()
        return jsonify({'logged_in': False, 'pin_verified': False})
    return jsonify({
        'logged_in': True,
        'pin_verified': session.get('pin_verified', False),
        'user_id': uid,
        'name': user['name'],
        'role': user.get('role', 'user'),
        'lock_timeout': user.get('lock_timeout', 15)
    })


@app.route('/api/auth/lock', methods=['POST'])
@require_login
def api_lock():
    session['pin_verified'] = False
    return jsonify({'ok': True})

# ══════════════════════════════════════════════════════════════════════════════
# PROFILE APIs
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/profile', methods=['GET'])
@require_login
@require_pin_verified
def api_get_profile():
    user = get_current_user()
    return jsonify({
        'ok': True,
        'name': user['name'],
        'block': user.get('block', ''),
        'email': user['email'],
        'mobile': user['mobile'],
        'role': user.get('role', 'user'),
        'lock_timeout': user.get('lock_timeout', 15)
    })


@app.route('/api/profile/lock-timeout', methods=['POST'])
@require_login
@require_pin_verified
def api_set_lock_timeout():
    body    = request.get_json(force=True)
    timeout = body.get('timeout')
    if timeout not in [15, 60, 1440]:
        return jsonify({'ok': False, 'msg': 'Invalid timeout value!'}), 400

    users = load_users()
    uid   = current_user_id()
    users[uid]['lock_timeout'] = timeout
    save_users(users)
    return jsonify({'ok': True, 'msg': 'Lock timeout update ho gaya!', 'lock_timeout': timeout})


@app.route('/api/profile/change-pin', methods=['POST'])
@require_login
@require_pin_verified
def api_change_pin():
    body     = request.get_json(force=True)
    old_pin  = str(body.get('old_pin', ''))
    new_pin  = str(body.get('new_pin', ''))
    conf_pin = str(body.get('confirm_pin', ''))

    users = load_users()
    uid   = current_user_id()
    user  = users.get(uid)

    if hash_pin(old_pin) != user.get('pin_hash', ''):
        return jsonify({'ok': False, 'msg': 'Purana PIN galat hai!'}), 403
    if not new_pin.isdigit() or len(new_pin) != 4:
        return jsonify({'ok': False, 'msg': 'Naya PIN sirf 4 digit ka hona chahiye!'}), 400
    if new_pin != conf_pin:
        return jsonify({'ok': False, 'msg': 'Dono PIN match nahi karte!'}), 400

    users[uid]['pin_hash'] = hash_pin(new_pin)
    save_users(users)
    return jsonify({'ok': True, 'msg': 'PIN successfully change ho gaya!'})

# ══════════════════════════════════════════════════════════════════════════════
# DATA APIs (per-user)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/data', methods=['GET'])
@require_login
@require_pin_verified
def api_get_data():
    users = load_users()
    uid   = current_user_id()
    return jsonify(users[uid].get('data', get_default_data()))


@app.route('/api/data', methods=['POST'])
@require_login
@require_pin_verified
def api_save_data():
    body  = request.get_json(force=True)
    users = load_users()
    uid   = current_user_id()
    users[uid]['data'] = body
    save_users(users)
    return jsonify({'ok': True})

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT API
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/export/excel', methods=['GET'])
@require_login
@require_pin_verified
def api_export_excel():
    users = load_users()
    uid   = current_user_id()
    data  = users[uid].get('data', get_default_data())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR_FILL  = PatternFill("solid", fgColor="3730A3")
    HDR_FONT  = Font(name='Calibri', bold=True, color="FFFFFF", size=11)
    TTL_FILL  = PatternFill("solid", fgColor="4F46E5")
    TTL_FONT  = Font(name='Calibri', bold=True, color="FFFFFF", size=13)
    INFO_FILL = PatternFill("solid", fgColor="EEF2FF")
    INFO_FONT = Font(name='Calibri', italic=True, color="4338CA", size=9)
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    GRAY_FILL  = PatternFill("solid", fgColor="F9FAFB")
    PASS_BG   = PatternFill("solid", fgColor="F5F3FF")
    PASS_FONT = Font(name='Consolas', color="4338CA", size=10)
    CELL_FONT = Font(name='Calibri', color="111827", size=10)
    BOLD_FONT = Font(name='Calibri', color="111827", size=10, bold=True)
    thin   = Side(style='thin',   color="E5E7EB")
    medium = Side(style='medium', color="C7D2FE")
    BORDER     = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
    HDR_BORDER = Border(left=medium, right=medium, top=medium, bottom=medium)
    CENTER = Alignment(horizontal='center', vertical='center')
    LEFT   = Alignment(horizontal='left',   vertical='center')

    for cat in data.get('categories', []):
        ws      = wb.create_sheet(title=cat['name'][:31])
        columns = cat.get('columns', [])
        entries = cat.get('entries', [])
        n_cols  = max(len(columns), 1)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        tc = ws.cell(row=1, column=1, value=f"Password Manager  —  {cat['name'].upper()}")
        tc.fill = TTL_FILL; tc.font = TTL_FONT; tc.alignment = CENTER
        ws.row_dimensions[1].height = 32

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        ic = ws.cell(row=2, column=1,
            value=(f"Exported: {datetime.now().strftime('%d %B %Y  %I:%M %p')}"
                   f"   |   Total Entries: {len(entries)}"))
        ic.fill = INFO_FILL; ic.font = INFO_FONT; ic.alignment = CENTER
        ws.row_dimensions[2].height = 18

        HDR_ROW = 3
        for j, col in enumerate(columns, start=1):
            c = ws.cell(row=HDR_ROW, column=j, value=col.upper())
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = CENTER; c.border = HDR_BORDER
        ws.row_dimensions[HDR_ROW].height = 26

        for i, entry in enumerate(entries):
            row_num  = HDR_ROW + 1 + i
            row_fill = WHITE_FILL if i % 2 == 0 else GRAY_FILL
            for j in range(len(columns)):
                val     = entry[j] if j < len(entry) else ''
                is_pass = any(k in columns[j].lower() for k in ('pass','pwd','secret'))
                c = ws.cell(row=row_num, column=j + 1, value=val)
                c.fill = PASS_BG if is_pass else row_fill
                c.font = PASS_FONT if is_pass else CELL_FONT
                c.alignment = CENTER if is_pass else LEFT
                c.border = BORDER
            ws.row_dimensions[row_num].height = 20

        for j, col in enumerate(columns, start=1):
            col_vals = [str(col)] + [str(e[j-1]) if j-1 < len(e) else '' for e in entries]
            best = max(len(v) for v in col_vals)
            ws.column_dimensions[get_column_letter(j)].width = min(max(best+3, 12), 45)

        ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)
        if columns:
            ws.auto_filter.ref = (
                f"A{HDR_ROW}:{get_column_letter(len(columns))}{HDR_ROW + len(entries)}"
            )

    # Summary sheet
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.column_dimensions['A'].width = 28
    ws_sum.column_dimensions['B'].width = 16
    ws_sum.column_dimensions['C'].width = 45
    ws_sum.merge_cells('A1:C1')
    t = ws_sum.cell(row=1, column=1, value="Password Manager — Export Summary")
    t.fill = TTL_FILL; t.font = TTL_FONT; t.alignment = CENTER
    ws_sum.row_dimensions[1].height = 32
    ws_sum.merge_cells('A2:C2')
    d = ws_sum.cell(row=2, column=1,
        value=f"Generated: {datetime.now().strftime('%d %B %Y  %I:%M %p')}")
    d.fill = INFO_FILL; d.font = INFO_FONT; d.alignment = CENTER
    ws_sum.row_dimensions[2].height = 18
    for hdr, col in [("Category",1),("Total Entries",2),("Columns",3)]:
        c = ws_sum.cell(row=3, column=col, value=hdr.upper())
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = CENTER; c.border = HDR_BORDER
    ws_sum.row_dimensions[3].height = 26
    for i, cat in enumerate(data.get('categories', []), start=4):
        fill = WHITE_FILL if i % 2 == 0 else GRAY_FILL
        c1 = ws_sum.cell(row=i, column=1, value=cat['name'])
        c1.font = BOLD_FONT; c1.fill = fill; c1.alignment = LEFT; c1.border = BORDER
        c2 = ws_sum.cell(row=i, column=2, value=len(cat.get('entries',[])))
        c2.font = CELL_FONT; c2.fill = fill; c2.alignment = CENTER; c2.border = BORDER
        c3 = ws_sum.cell(row=i, column=3, value=", ".join(cat.get('columns',[])))
        c3.font = CELL_FONT; c3.fill = fill; c3.alignment = LEFT; c3.border = BORDER
        ws_sum.row_dimensions[i].height = 20

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"passwords_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ══════════════════════════════════════════════════════════════════════════════
# ADMIN APIs
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/admin/users', methods=['GET'])
@require_login
@require_pin_verified
@require_admin
def api_admin_users():
    users = load_users()
    result = []
    for uid, u in users.items():
        result.append({
            'id': uid,
            'name': u.get('name',''),
            'block': u.get('block',''),
            'email': u.get('email',''),
            'mobile': u.get('mobile',''),
            'role': u.get('role','user'),
            'lock_timeout': u.get('lock_timeout', 15)
        })
    return jsonify({'ok': True, 'users': result})


@app.route('/api/admin/reset-pin', methods=['POST'])
@require_login
@require_pin_verified
@require_admin
def api_admin_reset_pin():
    body    = request.get_json(force=True)
    uid     = str(body.get('user_id', ''))
    new_pin = str(body.get('new_pin', '1234'))

    if not new_pin.isdigit() or len(new_pin) != 4:
        return jsonify({'ok': False, 'msg': 'PIN sirf 4 digit ka hona chahiye!'}), 400

    users = load_users()
    if uid not in users:
        return jsonify({'ok': False, 'msg': 'User nahi mila!'}), 404
    if uid == current_user_id():
        return jsonify({'ok': False, 'msg': 'Apna PIN yahan se reset nahi kar sakte!'}), 400

    users[uid]['pin_hash'] = hash_pin(new_pin)
    save_users(users)
    return jsonify({'ok': True, 'msg': f'{users[uid]["name"]} ka PIN reset ho gaya! Naya PIN: {new_pin}'})


@app.route('/api/admin/delete-user', methods=['POST'])
@require_login
@require_pin_verified
@require_admin
def api_admin_delete_user():
    body = request.get_json(force=True)
    uid  = str(body.get('user_id', ''))

    if uid == current_user_id():
        return jsonify({'ok': False, 'msg': 'Apna khud ka account delete nahi kar sakte!'}), 400

    users = load_users()
    if uid not in users:
        return jsonify({'ok': False, 'msg': 'User nahi mila!'}), 404

    name = users[uid].get('name', uid)
    del users[uid]
    save_users(users)
    return jsonify({'ok': True, 'msg': f'{name} ka account delete ho gaya!'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=7730, debug=False)
